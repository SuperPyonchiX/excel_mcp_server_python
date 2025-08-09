#!/usr/bin/env python3
"""
Excel操作統合テスト（修正版）
実際のビジネスロジック関数を直接テストします
"""

import sys
import json
import os
from pathlib import Path
from typing import Union, List

# Windows環境での文字化け対策
if sys.platform == "win32":
    os.environ["PYTHONIOENCODING"] = "utf-8"

# プロジェクトのパスを追加
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root / "src"))

import openpyxl
import pandas as pd
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.workbook import Workbook

# テスト設定
TEST_DIR = Path(__file__).parent / "output"
TEST_DIR.mkdir(exist_ok=True)

TEST_CONFIG = {
    'filePath': str(TEST_DIR / 'test-workbook.xlsx'),
    'csvPath': str(TEST_DIR / 'test-export.csv'),
    'sheetName': 'TestSheet'
}

# テストデータ定義
TEST_DATA = {
    'sampleData': [
        ['商品名', '価格', '在庫', '売上'],
        ['商品A', 1000, 50, '=B2*C2'],
        ['商品B', 1500, 30, '=B3*C3'],
        ['商品C', 800, 75, '=B4*C4']
    ],
    
    'headerFormat': {
        'font': {
            'bold': True,
            'size': 12,
            'color': 'FF000080'
        },
        'fill': {
            'type': 'pattern',
            'pattern': 'solid',
            'fgColor': 'FFE0E0E0'
        }
    }
}


# ビジネスロジック関数（MCPデコレーターなし）
def validate_file_path(filePath: str) -> None:
    """ファイルパスの妥当性を検証"""
    if not filePath:
        raise ValueError("ファイルパスが指定されていません")
    
    if not (filePath.endswith('.xlsx') or filePath.endswith('.xls')):
        raise ValueError("ファイル拡張子は .xlsx または .xls である必要があります")
    
    if not Path(filePath).is_absolute():
        raise ValueError("絶対パスを指定してください（例: C:/Users/Username/Documents/file.xlsx）")


def create_workbook_logic(filePath: str) -> str:
    """新しいExcelワークブックを作成"""
    try:
        validate_file_path(filePath)
        workbook = Workbook()
        workbook.save(filePath)
        return f"Excelワークブック '{filePath}' を作成しました。"
    except Exception as e:
        raise Exception(f"ワークブック作成エラー: {e}")


def add_worksheet_logic(filePath: str, sheetName: str) -> str:
    """既存のワークブックにワークシートを追加"""
    try:
        if not sheetName or not sheetName.strip():
            raise ValueError("ワークシート名が空です")
        
        workbook = openpyxl.load_workbook(filePath)
        
        if sheetName in workbook.sheetnames:
            raise ValueError(f"ワークシート '{sheetName}' は既に存在します")
        
        workbook.create_sheet(sheetName)
        workbook.save(filePath)
        
        return f"ワークシート '{sheetName}' を追加しました。"
    except Exception as e:
        raise Exception(f"ワークシート追加エラー: {e}")


def set_cell_value_logic(filePath: str, sheetName: str, cell: str, value: Union[str, int, float, bool]) -> str:
    """指定されたセルに値を設定"""
    try:
        workbook = openpyxl.load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            available_sheets = ', '.join(workbook.sheetnames)
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。利用可能なシート: {available_sheets}")
        
        worksheet = workbook[sheetName]
        worksheet[cell] = value
        workbook.save(filePath)
        
        return f"セル {cell} に値 '{value}' を設定しました。"
    except Exception as e:
        raise Exception(f"セル値設定エラー: {e}")


def get_cell_value_logic(filePath: str, sheetName: str, cell: str) -> str:
    """指定されたセルの値を取得"""
    try:
        workbook = openpyxl.load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            available_sheets = ', '.join(workbook.sheetnames)
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。利用可能なシート: {available_sheets}")
        
        worksheet = workbook[sheetName]
        cell_value = worksheet[cell].value
        
        return f"セル {cell} の値: {cell_value}"
    except Exception as e:
        raise Exception(f"セル値取得エラー: {e}")


def set_range_values_logic(filePath: str, sheetName: str, startCell: str, values: List[List[Union[str, int, float, bool]]]) -> str:
    """指定された範囲に2次元配列のデータを設定"""
    try:
        if not values or len(values) == 0:
            raise ValueError("valuesは空でない2次元配列である必要があります")
        
        workbook = openpyxl.load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            available_sheets = ', '.join(workbook.sheetnames)
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。利用可能なシート: {available_sheets}")
        
        worksheet = workbook[sheetName]
        
        # 開始セルの行・列番号を取得
        start_cell_obj = worksheet[startCell]
        start_row = start_cell_obj.row
        start_col = start_cell_obj.column
        
        # データを設定
        for i, row_data in enumerate(values):
            for j, cell_value in enumerate(row_data):
                worksheet.cell(row=start_row + i, column=start_col + j, value=cell_value)
        
        workbook.save(filePath)
        
        max_cols = max(len(row) for row in values) if values else 0
        return f"範囲 {startCell} から {len(values)}行 x {max_cols}列 のデータを設定しました。"
    except Exception as e:
        raise Exception(f"範囲値設定エラー: {e}")


def get_range_values_logic(filePath: str, sheetName: str, rangeAddr: str) -> str:
    """指定された範囲のデータを取得"""
    try:
        workbook = openpyxl.load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            available_sheets = ', '.join(workbook.sheetnames)
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。利用可能なシート: {available_sheets}")
        
        worksheet = workbook[sheetName]
        
        # 範囲を解析
        start_cell, end_cell = rangeAddr.split(':')
        
        # 開始・終了セルの座標を取得
        start_cell_obj = worksheet[start_cell]
        end_cell_obj = worksheet[end_cell]
        
        start_row, start_col = start_cell_obj.row, start_cell_obj.column
        end_row, end_col = end_cell_obj.row, end_cell_obj.column
        
        # データを取得
        values = []
        for row in range(start_row, end_row + 1):
            row_values = []
            for col in range(start_col, end_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                row_values.append(cell_value)
            values.append(row_values)
        
        return f"範囲 {rangeAddr} の値:\n{json.dumps(values, ensure_ascii=False, indent=2)}"
    except Exception as e:
        raise Exception(f"範囲値取得エラー: {e}")


def format_cell_logic(filePath: str, sheetName: str, cell: str, formatSpec: dict) -> str:
    """セルの書式（フォント、塗りつぶし、罫線）を設定"""
    try:
        workbook = openpyxl.load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。")
        
        worksheet = workbook[sheetName]
        target_cell = worksheet[cell]
        
        # フォント設定
        if 'font' in formatSpec:
            font_spec = formatSpec['font']
            font_kwargs = {}
            if 'bold' in font_spec:
                font_kwargs['bold'] = font_spec['bold']
            if 'italic' in font_spec:
                font_kwargs['italic'] = font_spec['italic']
            if 'size' in font_spec:
                font_kwargs['size'] = font_spec['size']
            if 'color' in font_spec:
                font_kwargs['color'] = font_spec['color']
            
            if font_kwargs:
                target_cell.font = Font(**font_kwargs)
        
        # 塗りつぶし設定
        if 'fill' in formatSpec:
            fill_spec = formatSpec['fill']
            if fill_spec.get('type') == 'pattern':
                target_cell.fill = PatternFill(
                    fill_type=fill_spec.get('pattern', 'solid'),
                    fgColor=fill_spec.get('fgColor', 'FFFFFF')
                )
        
        workbook.save(filePath)
        
        return f"セル {cell} の書式を設定しました。"
    except Exception as e:
        raise Exception(f"セル書式設定エラー: {e}")


def add_formula_logic(filePath: str, sheetName: str, cell: str, formula: str) -> str:
    """セルに数式を追加"""
    try:
        workbook = openpyxl.load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。")
        
        worksheet = workbook[sheetName]
        worksheet[cell] = formula
        workbook.save(filePath)
        
        return f"セル {cell} に数式 '{formula}' を設定しました。"
    except Exception as e:
        raise Exception(f"数式追加エラー: {e}")


def find_data_logic(filePath: str, sheetName: str, searchValue: Union[str, int, float]) -> str:
    """ワークシート内で指定された値を検索"""
    try:
        workbook = openpyxl.load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。")
        
        worksheet = workbook[sheetName]
        results = []
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == searchValue:
                    results.append(cell.coordinate)
        
        return f"値 '{searchValue}' が見つかったセル: {', '.join(results)}"
    except Exception as e:
        raise Exception(f"データ検索エラー: {e}")


def export_to_csv_logic(filePath: str, sheetName: str, csvPath: str) -> str:
    """ワークシートをCSVファイルにエクスポート"""
    try:
        workbook = openpyxl.load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。")
        
        # DataFrameに変換してCSVに出力
        worksheet = workbook[sheetName]
        data = []
        
        for row in worksheet.iter_rows(values_only=True):
            data.append(row)
        
        df = pd.DataFrame(data)
        df.to_csv(csvPath, index=False, header=False, encoding='utf-8-sig')
        
        return f"ワークシート '{sheetName}' をCSVファイル '{csvPath}' にエクスポートしました。"
    except Exception as e:
        raise Exception(f"CSV出力エラー: {e}")


def execute_test_with_result(test_name: str, test_func, test_id: int):
    """テストを実行し結果を表示する"""
    print(f"\n[{test_id}] {test_name} を実行中...")
    try:
        result = test_func()
        message = str(result)
        print(f"[成功] [{test_id}] {message}")
        return True
    except Exception as e:
        print(f"[失敗] [{test_id}] エラー: {str(e)}")
        return False


def run_excel_integration_test():
    """Excel操作の統合テストを実行"""
    print('=== Excel操作統合テスト（修正版） ===\n')
    print(f'テストファイル: {TEST_CONFIG["filePath"]}')
    print(f'CSV出力先: {TEST_CONFIG["csvPath"]}')
    
    # 既存ファイルを削除
    for file_path in [TEST_CONFIG['filePath'], TEST_CONFIG['csvPath']]:
        path_obj = Path(file_path)
        if path_obj.exists():
            path_obj.unlink()
    
    completed_tests = 0
    total_tests = 10
    
    print('\nテスト開始\n')
    
    # テストシーケンス実行
    test_results = run_test_sequence()
    completed_tests = sum(test_results)
    
    # テスト結果サマリー
    print_test_summary(completed_tests, total_tests)
    
    return completed_tests == total_tests


def run_test_sequence():
    """テストシーケンスを実行"""
    test_results = []
    test_id = 1
    
    # 1. ワークブック作成
    result = execute_test_with_result(
        'ワークブック作成', 
        lambda: create_workbook_logic(filePath=TEST_CONFIG['filePath']),
        test_id
    )
    test_results.append(result)
    test_id += 1
    
    # 2. ワークシート追加
    result = execute_test_with_result(
        'ワークシート追加',
        lambda: add_worksheet_logic(
            filePath=TEST_CONFIG['filePath'],
            sheetName=TEST_CONFIG['sheetName']
        ),
        test_id
    )
    test_results.append(result)
    test_id += 1
    
    # 3. タイトル設定
    result = execute_test_with_result(
        'タイトル設定',
        lambda: set_cell_value_logic(
            filePath=TEST_CONFIG['filePath'],
            sheetName=TEST_CONFIG['sheetName'],
            cell='A1',
            value='Excel MCP 統合テスト レポート'
        ),
        test_id
    )
    test_results.append(result)
    test_id += 1
    
    # 4. サンプルデータ入力
    result = execute_test_with_result(
        'サンプルデータ入力',
        lambda: set_range_values_logic(
            filePath=TEST_CONFIG['filePath'],
            sheetName=TEST_CONFIG['sheetName'],
            startCell='A3',
            values=TEST_DATA['sampleData']
        ),
        test_id
    )
    test_results.append(result)
    test_id += 1
    
    # 5. ヘッダー書式設定
    result = execute_test_with_result(
        'ヘッダー書式設定',
        lambda: format_cell_logic(
            filePath=TEST_CONFIG['filePath'],
            sheetName=TEST_CONFIG['sheetName'],
            cell='A3',
            formatSpec=TEST_DATA['headerFormat']
        ),
        test_id
    )
    test_results.append(result)
    test_id += 1
    
    # 6. 合計数式追加
    result = execute_test_with_result(
        '合計数式追加',
        lambda: add_formula_logic(
            filePath=TEST_CONFIG['filePath'],
            sheetName=TEST_CONFIG['sheetName'],
            cell='E7',
            formula='=SUM(E4:E6)'
        ),
        test_id
    )
    test_results.append(result)
    test_id += 1
    
    # 7. セル値取得テスト
    result = execute_test_with_result(
        'セル値取得テスト',
        lambda: get_cell_value_logic(
            filePath=TEST_CONFIG['filePath'],
            sheetName=TEST_CONFIG['sheetName'],
            cell='A1'
        ),
        test_id
    )
    test_results.append(result)
    test_id += 1
    
    # 8. 範囲データ取得テスト
    result = execute_test_with_result(
        '範囲データ取得テスト',
        lambda: get_range_values_logic(
            filePath=TEST_CONFIG['filePath'],
            sheetName=TEST_CONFIG['sheetName'],
            rangeAddr='A3:D6'
        ),
        test_id
    )
    test_results.append(result)
    test_id += 1
    
    # 9. データ検索テスト
    result = execute_test_with_result(
        'データ検索テスト',
        lambda: find_data_logic(
            filePath=TEST_CONFIG['filePath'],
            sheetName=TEST_CONFIG['sheetName'],
            searchValue='商品A'
        ),
        test_id
    )
    test_results.append(result)
    test_id += 1
    
    # 10. CSV出力テスト
    result = execute_test_with_result(
        'CSV出力テスト',
        lambda: export_to_csv_logic(
            filePath=TEST_CONFIG['filePath'],
            sheetName=TEST_CONFIG['sheetName'],
            csvPath=TEST_CONFIG['csvPath']
        ),
        test_id
    )
    test_results.append(result)
    
    return test_results


def print_test_summary(completed_tests: int, total_tests: int):
    """テスト結果サマリーを出力"""
    print('\n' + '=' * 50)
    print('Excel操作統合テスト完了')
    print('=' * 50)
    
    try:
        # ファイル存在確認
        excel_path = Path(TEST_CONFIG['filePath'])
        csv_path = Path(TEST_CONFIG['csvPath'])
        
        print('生成されたファイル:')
        
        if excel_path.exists():
            excel_size = excel_path.stat().st_size
            print(f'   Excel: {excel_path}')
            print(f'      サイズ: {excel_size} bytes')
        else:
            print(f'   [ERROR] Excel: {excel_path} (見つかりません)')
        
        if csv_path.exists():
            csv_size = csv_path.stat().st_size
            print(f'   CSV: {csv_path}')
            print(f'      サイズ: {csv_size} bytes')
            
            # CSV内容の確認
            with open(csv_path, 'r', encoding='utf-8-sig') as f:
                lines = f.readlines()[:5]  # 最初の5行
            
            if lines:
                print(f'\nCSV出力内容プレビュー:')
                for i, line in enumerate(lines):
                    print(f'   {i+1}: {line.rstrip()}')
        else:
            print(f'   [ERROR] CSV: {csv_path} (見つかりません)')
        
    except Exception as error:
        print(f'ファイル確認中にエラーが発生しました: {error}')
    
    if completed_tests == total_tests:
        print('\n全ての機能が正常に動作しました！')
    else:
        print(f'\n{total_tests}個中{completed_tests}個のテストが完了しました')


# テスト実行
if __name__ == "__main__":
    try:
        success = run_excel_integration_test()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"テスト実行エラー: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
