#!/usr/bin/env python3
"""
Excel MCP Server - Python版 (FastMCP使用)
AIエージェントがExcelを自由に操作できるModel Context Protocol (MCP) サーバーです。
"""

import json
import os
import re
from pathlib import Path
from typing import Any, List, Optional, Union

import openpyxl
import pandas as pd
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.workbook import Workbook

from mcp.server.fastmcp import FastMCP

# FastMCPサーバーインスタンスを作成
mcp = FastMCP("Excel MCP Server")

# ワークブックキャッシュ
workbook_cache = {}
open_workbook_paths = set()


def validate_file_path(filePath: str) -> None:
    """ファイルパスの妥当性を検証"""
    if not filePath:
        raise ValueError("ファイルパスが指定されていません")
    
    if not (filePath.endswith('.xlsx') or filePath.endswith('.xls')):
        raise ValueError("ファイル拡張子は .xlsx または .xls である必要があります")
    
    if not os.path.isabs(filePath):
        raise ValueError("絶対パスを指定してください（例: C:/Users/Username/Documents/file.xlsx）")


def validate_cell_address(cell: str) -> None:
    """セルアドレスの妥当性を検証"""
    pattern = r'^[A-Z]+[1-9]\d*$'
    if not re.match(pattern, cell):
        raise ValueError(f"無効なセル位置: '{cell}'。正しい形式: A1, B2, AA10など")


def validate_range_address(rangeAddr: str) -> None:
    """範囲アドレスの妥当性を検証"""
    pattern = r'^[A-Z]+[1-9]\d*:[A-Z]+[1-9]\d*$'
    if not re.match(pattern, rangeAddr):
        raise ValueError(f"無効な範囲指定: '{rangeAddr}'。正しい形式: A1:C3, B2:D10など")


def get_sheet_names(workbook: Workbook) -> str:
    """ワークブック内のシート名一覧を取得"""
    return ', '.join(workbook.sheetnames)


def load_workbook(filePath: str) -> Workbook:
    """ワークブックを読み込む"""
    validate_file_path(filePath)
    return openpyxl.load_workbook(filePath)


@mcp.tool()
def create_workbook(filePath: str) -> str:
    """
    新しいExcelワークブックを作成します
    
    Args:
        filePath: 作成するExcelファイルの絶対パス（例: C:/Users/Username/Documents/report.xlsx）。ファイル拡張子は.xlsxである必要があります
    """
    try:
        validate_file_path(filePath)
        
        workbook = Workbook()
        workbook.save(filePath)
        
        return f"Excelワークブック '{filePath}' を作成しました。"
    except Exception as e:
        raise Exception(f"ワークブック作成エラー: {e}")


@mcp.tool()
def open_workbook(filePath: str) -> str:
    """
    既存のExcelワークブックを開いて情報を表示します
    
    Args:
        filePath: 開くExcelファイルの絶対パス。既存のファイルである必要があります
    """
    try:
        validate_file_path(filePath)
        
        if not os.path.exists(filePath):
            raise FileNotFoundError(f"ファイルが見つかりません: {filePath}")
        
        workbook = openpyxl.load_workbook(filePath)
        
        # キャッシュに保存
        workbook_cache[filePath] = workbook
        open_workbook_paths.add(filePath)
        
        sheet_names = get_sheet_names(workbook)
        sheet_count = len(workbook.sheetnames)
        
        return f"Excelワークブック '{filePath}' を開きました。\nワークシート数: {sheet_count}\nシート名: {sheet_names}"
    except Exception as e:
        raise Exception(f"ワークブック読み込みエラー: {e}")


@mcp.tool()
def get_workbook_info(filePath: str) -> str:
    """
    Excelワークブックの詳細情報を取得します（シート一覧、メタデータ等）
    
    Args:
        filePath: 情報を取得するExcelファイルの絶対パス
    """
    try:
        validate_file_path(filePath)
        
        if not os.path.exists(filePath):
            raise FileNotFoundError(f"ファイルが見つかりません: {filePath}")
        
        workbook = openpyxl.load_workbook(filePath)
        
        # ファイル情報を取得
        file_stat = os.stat(filePath)
        
        info = {
            "ファイルパス": filePath,
            "ワークシート数": len(workbook.sheetnames),
            "ワークシート名一覧": workbook.sheetnames,
            "ファイルサイズ": f"{file_stat.st_size} bytes",
            "最終更新日時": pd.Timestamp.fromtimestamp(file_stat.st_mtime).isoformat()
        }
        
        return f"ワークブック情報:\n{json.dumps(info, ensure_ascii=False, indent=2)}"
    except Exception as e:
        raise Exception(f"ワークブック情報取得エラー: {e}")


@mcp.tool()
def add_worksheet(filePath: str, sheetName: str) -> str:
    """
    既存のワークブックにワークシートを追加します
    
    Args:
        filePath: 対象のExcelファイルの絶対パス。既存のファイルである必要があります
        sheetName: 作成するワークシート名。英数字、日本語、アンダースコア、ハイフンが使用可能です
    """
    try:
        if not sheetName or not sheetName.strip():
            raise ValueError("ワークシート名が空です")
        
        workbook = load_workbook(filePath)
        
        if sheetName in workbook.sheetnames:
            raise ValueError(f"ワークシート '{sheetName}' は既に存在します")
        
        workbook.create_sheet(sheetName)
        workbook.save(filePath)
        
        return f"ワークシート '{sheetName}' を追加しました。"
    except Exception as e:
        raise Exception(f"ワークシート追加エラー: {e}")


@mcp.tool()
def set_cell_value(filePath: str, sheetName: str, cell: str, value: Union[str, int, float, bool]) -> str:
    """
    指定されたセルに値を設定します
    
    Args:
        filePath: 対象のExcelファイルの絶対パス
        sheetName: 対象のワークシート名。既存のワークシート名を指定してください
        cell: セル位置。A1形式で指定（例: A1, B2, AA10, Z99）。範囲指定（A1:B2）は不可
        value: セルに設定する値。文字列、数値、真偽値のいずれか
    """
    try:
        validate_cell_address(cell)
        
        workbook = load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            available_sheets = get_sheet_names(workbook)
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。利用可能なシート: {available_sheets}")
        
        worksheet = workbook[sheetName]
        worksheet[cell] = value
        workbook.save(filePath)
        
        return f"セル {cell} に値 '{value}' を設定しました。"
    except Exception as e:
        raise Exception(f"セル値設定エラー: {e}")


@mcp.tool()
def get_cell_value(filePath: str, sheetName: str, cell: str) -> str:
    """
    指定されたセルの値を取得します
    
    Args:
        filePath: 対象のExcelファイルの絶対パス
        sheetName: 対象のワークシート名
        cell: セル位置。A1形式で指定（例: A1, B2, AA10）
    """
    try:
        validate_cell_address(cell)
        
        workbook = load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            available_sheets = get_sheet_names(workbook)
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。利用可能なシート: {available_sheets}")
        
        worksheet = workbook[sheetName]
        cell_value = worksheet[cell].value
        
        return f"セル {cell} の値: {cell_value}"
    except Exception as e:
        raise Exception(f"セル値取得エラー: {e}")


@mcp.tool()
def set_range_values(filePath: str, sheetName: str, startCell: str, values: List[List[Union[str, int, float, bool]]]) -> str:
    """
    指定された範囲に2次元配列のデータを設定します
    
    Args:
        filePath: 対象のExcelファイルの絶対パス
        sheetName: 対象のワークシート名
        startCell: データ入力を開始するセル位置（例: A1）。ここから右下方向にデータが入力されます
        values: 2次元配列のデータ。外側の配列が行、内側の配列が列を表します。例: [["商品名", "価格"], ["商品A", 1000]]
    """
    try:
        validate_cell_address(startCell)
        
        if not values or len(values) == 0:
            raise ValueError("valuesは空でない2次元配列である必要があります")
        
        # 2次元配列の検証
        for i, row in enumerate(values):
            if not isinstance(row, list):
                raise ValueError(f"{i+1}行目が配列ではありません。2次元配列を指定してください")
        
        workbook = load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            available_sheets = get_sheet_names(workbook)
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。利用可能なシート: {available_sheets}")
        
        worksheet = workbook[sheetName]
        
        # 開始セルの行・列番号を取得
        start_cell_obj = worksheet[startCell]
        start_row = start_cell_obj.row
        start_col = start_cell_obj.column
        
        # データを設定
        for i, row_data in enumerate(values):
            for j, cell_value in enumerate(row_data):
                worksheet.cell(row=start_row + i, column=start_col + j, value=cell_value)
        
        workbook.save(filePath)
        
        max_cols = max(len(row) for row in values) if values else 0
        return f"範囲 {startCell} から {len(values)}行 x {max_cols}列 のデータを設定しました。"
    except Exception as e:
        raise Exception(f"範囲値設定エラー: {e}")


@mcp.tool()
def get_range_values(filePath: str, sheetName: str, rangeAddr: str) -> str:
    """
    指定された範囲のデータを取得します
    
    Args:
        filePath: 対象のExcelファイルの絶対パス
        sheetName: 対象のワークシート名
        rangeAddr: 取得する範囲。A1:C3形式で指定（例: A1:C10, B2:D5）
    """
    try:
        validate_range_address(rangeAddr)
        
        workbook = load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            available_sheets = get_sheet_names(workbook)
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。利用可能なシート: {available_sheets}")
        
        worksheet = workbook[sheetName]
        
        # 範囲を解析
        start_cell, end_cell = rangeAddr.split(':')
        
        # 開始・終了セルの座標を取得
        start_cell_obj = worksheet[start_cell]
        end_cell_obj = worksheet[end_cell]
        
        start_row, start_col = start_cell_obj.row, start_cell_obj.column
        end_row, end_col = end_cell_obj.row, end_cell_obj.column
        
        # データを取得
        values = []
        for row in range(start_row, end_row + 1):
            row_values = []
            for col in range(start_col, end_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                row_values.append(cell_value)
            values.append(row_values)
        
        return f"範囲 {rangeAddr} の値:\n{json.dumps(values, ensure_ascii=False, indent=2)}"
    except Exception as e:
        raise Exception(f"範囲値取得エラー: {e}")


@mcp.tool()
def format_cell(filePath: str, sheetName: str, cell: str, formatSpec: dict) -> str:
    """
    セルの書式（フォント、塗りつぶし、罫線）を設定します
    
    Args:
        filePath: Excelファイルのパス
        sheetName: ワークシート名
        cell: セル位置（例: A1）
        formatSpec: セルの書式設定（辞書形式）
            - font: フォント設定 {"bold": bool, "italic": bool, "size": int, "color": str}
            - fill: 塗りつぶし設定 {"type": "pattern", "pattern": str, "fgColor": str}
            - border: 罫線設定 {"top": {"style": str, "color": str}, ...}
    """
    try:
        validate_cell_address(cell)
        
        workbook = load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。")
        
        worksheet = workbook[sheetName]
        target_cell = worksheet[cell]
        
        # フォント設定
        if 'font' in formatSpec:
            font_spec = formatSpec['font']
            font_kwargs = {}
            if 'bold' in font_spec:
                font_kwargs['bold'] = font_spec['bold']
            if 'italic' in font_spec:
                font_kwargs['italic'] = font_spec['italic']
            if 'size' in font_spec:
                font_kwargs['size'] = font_spec['size']
            if 'color' in font_spec:
                font_kwargs['color'] = font_spec['color']
            
            if font_kwargs:
                target_cell.font = Font(**font_kwargs)
        
        # 塗りつぶし設定
        if 'fill' in formatSpec:
            fill_spec = formatSpec['fill']
            if fill_spec.get('type') == 'pattern':
                target_cell.fill = PatternFill(
                    fill_type=fill_spec.get('pattern', 'solid'),
                    fgColor=fill_spec.get('fgColor', 'FFFFFF')
                )
        
        # 罫線設定
        if 'border' in formatSpec:
            border_spec = formatSpec['border']
            border_kwargs = {}
            for side_name in ['top', 'left', 'bottom', 'right']:
                if side_name in border_spec:
                    side_config = border_spec[side_name]
                    border_kwargs[side_name] = Side(
                        style=side_config.get('style', 'thin'),
                        color=side_config.get('color', '000000')
                    )
            
            if border_kwargs:
                target_cell.border = Border(**border_kwargs)
        
        workbook.save(filePath)
        
        return f"セル {cell} の書式を設定しました。"
    except Exception as e:
        raise Exception(f"セル書式設定エラー: {e}")


@mcp.tool()
def add_formula(filePath: str, sheetName: str, cell: str, formula: str) -> str:
    """
    セルに数式を追加します
    
    Args:
        filePath: Excelファイルのパス
        sheetName: ワークシート名
        cell: セル位置（例: A1）
        formula: 数式（=SUM(A1:A10)など、=で始まる）
    """
    try:
        validate_cell_address(cell)
        
        workbook = load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。")
        
        worksheet = workbook[sheetName]
        worksheet[cell] = formula
        workbook.save(filePath)
        
        return f"セル {cell} に数式 '{formula}' を設定しました。"
    except Exception as e:
        raise Exception(f"数式追加エラー: {e}")


@mcp.tool()
def find_data(filePath: str, sheetName: str, searchValue: Union[str, int, float]) -> str:
    """
    ワークシート内で指定された値を検索します
    
    Args:
        filePath: Excelファイルのパス
        sheetName: ワークシート名
        searchValue: 検索する値
    """
    try:
        workbook = load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。")
        
        worksheet = workbook[sheetName]
        results = []
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == searchValue:
                    results.append(cell.coordinate)
        
        return f"値 '{searchValue}' が見つかったセル: {', '.join(results)}"
    except Exception as e:
        raise Exception(f"データ検索エラー: {e}")


@mcp.tool()
def export_to_csv(filePath: str, sheetName: str, csvPath: str) -> str:
    """
    ワークシートをCSVファイルにエクスポートします
    
    Args:
        filePath: Excelファイルのパス（既存ファイル）
        sheetName: ワークシート名（既存シート）
        csvPath: CSVファイルの出力パス
    """
    try:
        workbook = load_workbook(filePath)
        
        if sheetName not in workbook.sheetnames:
            raise ValueError(f"ワークシート '{sheetName}' が見つかりません。")
        
        # DataFrameに変換してCSVに出力
        worksheet = workbook[sheetName]
        data = []
        
        for row in worksheet.iter_rows(values_only=True):
            data.append(row)
        
        df = pd.DataFrame(data)
        df.to_csv(csvPath, index=False, header=False, encoding='utf-8-sig')
        
        return f"ワークシート '{sheetName}' をCSVファイル '{csvPath}' にエクスポートしました。"
    except Exception as e:
        raise Exception(f"CSV出力エラー: {e}")


@mcp.tool()
def close_workbook(filePath: str) -> str:
    """
    開いているExcelワークブックを閉じてメモリから解放します
    
    Args:
        filePath: 閉じるExcelファイルの絶対パス。現在開いているファイルのパスを指定してください
    """
    try:
        validate_file_path(filePath)
        
        was_open = filePath in open_workbook_paths
        if was_open:
            if filePath in workbook_cache:
                del workbook_cache[filePath]
            open_workbook_paths.discard(filePath)
            return f"Excelワークブック '{filePath}' を閉じました。メモリから解放されました。"
        else:
            return f"Excelワークブック '{filePath}' は開かれていませんでした。"
    except Exception as e:
        raise Exception(f"ワークブック終了エラー: {e}")


@mcp.tool()
def list_open_workbooks() -> str:
    """
    現在開いているExcelワークブックの一覧を表示します
    """
    try:
        if len(open_workbook_paths) == 0:
            return "現在開いているワークブックはありません。"
        
        open_list = list(open_workbook_paths)
        info = {
            "開いているワークブック数": len(open_list),
            "ファイル一覧": open_list
        }
        
        return f"開いているワークブック:\n{json.dumps(info, ensure_ascii=False, indent=2)}"
    except Exception as e:
        raise Exception(f"開いているワークブック一覧取得エラー: {e}")


if __name__ == "__main__":
    mcp.run()
